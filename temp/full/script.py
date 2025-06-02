import openpyxl
from openpyxl.styles import Font
from faker import Faker
import random
import transliterate
import json
import re

# Инициализация Faker для русских данных
fake = Faker('ru_RU')

# Функция для загрузки групп из JSON файла
def load_groups_from_json(json_file):
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    unique_groups = set()
    for entry in data:
        group = entry.get('Группа', '')
        if group:
            unique_groups.add(group)
    
    return sorted(unique_groups)

# Функция для определения потока по группе
def get_stream(group):
    # Удаляем все не буквы и не цифры из названия группы
    clean_group = re.sub(r'[^\w]', '', group)
    
    # Ищем цифровую часть
    match = re.search(r'(\d+)', clean_group)
    if not match:
        return group
    
    numeric_part = match.group(1)
    prefix = clean_group[:match.start()]
    
    if len(numeric_part) == 3:
        return f"{prefix}{numeric_part[0]}**"
    elif len(numeric_part) == 2:
        return f"{prefix}{numeric_part[0]}*"
    else:
        return group

# Функция для генерации английского email на основе русского ФИО
def generate_english_email(last_name: str, first_name: str, middle_name: str) -> str:
    try:
        # Транслитерируем кириллицу в латиницу
        last_name_lat = transliterate.translit(last_name, reversed=True).replace("'", "")
        first_initial = transliterate.translit(first_name[0], reversed=True).replace("'", "") if first_name else ""
        middle_initial = transliterate.translit(middle_name[0], reversed=True).replace("'", "") if middle_name else ""
        
        # Формируем email
        emailFirst = Faker()
        emailSecond = Faker()
        emailThird = Faker()

        email = emailFirst.email() + emailSecond.email() + emailThird.email() 

        return email
    except:
        # Если транслитерация не сработала, используем случайный email
        emailFirst = Faker()
        emailSecond = Faker()
        emailThird = Faker()

        email = emailFirst.email() + emailSecond.email() + emailThird.email() 

        return email

# Основная функция
def generate_student_lists(json_file, output_filename):
    # Загружаем группы из JSON
    groups = load_groups_from_json(json_file)
    
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    
    # Удаляем дефолтный лист, если он есть
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    for group in groups:
        # Создаем лист для каждой группы
        ws = wb.create_sheet(title=group)
        
        # Заголовки
        headers = ["№", "Фамилия", "Имя", "Отчество", "Поток", "Группа", "Email"]
        ws.append(headers)
        
        # Жирный шрифт для заголовков
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Определяем поток по группе
        stream = get_stream(group)
        
        # Генерация 20 студентов для группы
        for i in range(1, 21):
            gender = random.choice(['male', 'female'])
            last_name = fake.last_name_male() if gender == 'male' else fake.last_name_female()
            first_name = fake.first_name_male() if gender == 'male' else fake.first_name_female()
            middle_name = fake.middle_name_male() if gender == 'male' else fake.middle_name_female()
            
            # Генерация английского email
            email = generate_english_email(last_name, first_name, middle_name)
            
            # Добавляем данные
            ws.append([
                i,
                last_name,
                first_name,
                middle_name,
                stream,  # Поток
                group,    # Группа
                email
            ])
        
        # Автоматическое выравнивание ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    wb.save(output_filename)
    print(f"Файл {output_filename} успешно создан!")

# Пример использования
if __name__ == "__main__":
    json_file = "fullSchedule.json"  # Укажите путь к вашему JSON-файлу
    output_filename = "fullStudentsExcel.xlsx"
    generate_student_lists(json_file, output_filename)