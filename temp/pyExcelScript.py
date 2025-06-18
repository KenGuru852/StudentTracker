import openpyxl
from openpyxl.styles import Font
from faker import Faker
import random
import transliterate

# Инициализация Faker для русских данных
fake = Faker('ru_RU')

# Создаем новую книгу Excel
wb = openpyxl.Workbook()

# Удаляем дефолтный лист, если он есть
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# Функция для генерации английского email на основе русского ФИО
def generate_english_email(last_name: str, first_name: str, middle_name: str) -> str:
    try:
        # Транслитерируем кириллицу в латиницу
        last_name_lat = transliterate.translit(last_name, reversed=True).replace("'", "")
        first_initial = transliterate.translit(first_name[0], reversed=True).replace("'", "") if first_name else ""
        middle_initial = transliterate.translit(middle_name[0], reversed=True).replace("'", "") if middle_name else ""
        
        # Формируем email
        email = f"{last_name_lat.lower()}.{first_initial.lower()}{middle_initial.lower()}@university.edu"
        return email
    except:
        # Если транслитерация не сработала, используем случайный email
        fake_en = Faker()
        return fake_en.email()

# Группы для потоков ИП-0** и ИП-1**
groups = (
    [f"ИП-01{i}" for i in range(1, 8)] +  # ИП-011, ИП-012, ..., ИП-017 (ИП-0**)
    [f"ИП-11{i}" for i in range(1, 8)]    # ИП-111, ИП-112, ..., ИП-117 (ИП-1**)
)

for group in groups:
    # Создаем лист для каждой группы
    ws = wb.create_sheet(title=group)
    
    # Заголовки
    headers = ["№", "Фамилия", "Имя", "Отчество", "Поток", "Группа", "Email", "СТ"]
    ws.append(headers)
    
    # Жирный шрифт для заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Определяем поток по группе
    if group.startswith("ИП-0"):
        stream = "ИП-0**"
    else:
        stream = "ИП-1**"
    
    # Генерация 20 студентов для группы
    students_data = []
    for i in range(1, 21):
        gender = random.choice(['male', 'female'])
        last_name = fake.last_name_male() if gender == 'male' else fake.last_name_female()
        first_name = fake.first_name_male() if gender == 'male' else fake.first_name_female()
        middle_name = fake.middle_name_male() if gender == 'male' else fake.middle_name_female()
        
        # Генерация английского email
        email = generate_english_email(last_name, first_name, middle_name)
        
        students_data.append([
            i,
            last_name,
            first_name,
            middle_name,
            stream,  # Поток
            group,   # Группа
            email,
            ""      # Пока пустое значение для СТ
        ])
    
    # Выбираем случайного студента для "+" в колонке СТ
    lucky_student_index = random.randint(0, 19)
    students_data[lucky_student_index][7] = "+"
    
    # Добавляем данные в таблицу
    for student in students_data:
        ws.append(student)
    
    # Автоматическое выравнивание ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

# Сохраняем файл
output_filename = "students_IP_groups.xlsx"
wb.save(output_filename)
print(f"Файл {output_filename} успешно создан!")